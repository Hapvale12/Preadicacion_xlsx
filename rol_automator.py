import pandas as pd
import re
import sys
import locale
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# --- 1. DATOS DE ENTRADA INTERACTIVA ---
print("----------------------------------------------------------")
print("  PROGRAMA DE ASIGNACIONES - EXTRACCIÓN DE TEXTO")
print("----------------------------------------------------------")
print("PASO 1: Copia el texto completo de las asignaciones de WhatsApp.")
print("PASO 2: Pégalo aquí (puedes pegar varias líneas a la vez).")
print("PASO 3: Una vez pegado, presiona Enter y luego Ctrl+Z (o Ctrl+D) y luego Enter de nuevo para finalizar la entrada.")
print("----------------------------------------------------------")

try:
    import sys
    texto_whatsapp = sys.stdin.read()
    if not texto_whatsapp.strip():
        raise EOFError 
except EOFError:
    print("\nNo se detectó entrada. Por favor, pega el texto de WhatsApp manualmente:")
    texto_whatsapp = input("Pega aquí el texto de WhatsApp: ")
    
# --- 2. EXTRACCIÓN Y TRANSFORMACIÓN DE DATOS ---
# REGEX PRINCIPAL: Busca DÍA y HORA, es flexible
REGEX_PATTERN = re.compile(
    r"^(?P<Dia>\w+)\s+"
    r"(?P<Hora>\d{1,2}:\d{2})\s*(?P<Periodo>am|pm)?\.?\s*" 
    r"(?P<Cuerpo>.+)$" 
)

datos_limpios = []
for linea in texto_whatsapp.strip().split('\n'):
    linea = linea.strip()
    if not linea.startswith(('Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo')):
        continue

    match = REGEX_PATTERN.search(linea)
    
    if match:
        data = match.groupdict()
        cuerpo = data['Cuerpo'].strip()
        
        final_pattern = r"(.*)\sTerritorio\s(\d+\s*.*?)[\.]?\s*(?P<Conductor>[\w\s\.]+)$"
        final_match = re.search(final_pattern, cuerpo)

        if final_match:
            # 1. El conductor es el último grupo
            conductor = final_match.group('Conductor').strip().strip('.')
            
            # 2. El Territorio es el segundo grupo
            territorio_bruto = final_match.group(2).strip().strip('.').replace(' y ', ', ')
            
            # 3. El Lugar es el primer grupo (ESTE ES EL VALOR QUE USAMOS DIRECTAMENTE)
            lugar_bruto = final_match.group(1).strip().strip('.').replace('Avenida', 'Av.')

            # Lógica de normalización de TERRITORIO
            normalized_pattern = r'(\d+\s*parte\s+[a-zA-Z])\s*(.*)'
            norm_match = re.search(normalized_pattern, territorio_bruto, flags=re.IGNORECASE)

            if norm_match:
                terr_base = norm_match.group(1).strip() 
                descripcion_adicional = norm_match.group(2).strip().strip('.') 
                
                terr_final = re.sub(r'\s*parte\s+([a-zA-Z]).*$', r'\1', terr_base, flags=re.IGNORECASE)
                terr_final = 'T.' + ''.join(terr_final.split()).upper()
                
                if descripcion_adicional:
                    # T.19A + Animas -> T.19AAnimas
                    desc_sufijo = "".join(descripcion_adicional.split()).capitalize()
                    terr_final += desc_sufijo
                
            else:
                # Territorio simple (Ej: T.18, T.3)
                terr_final = 'T.' + ''.join(territorio_bruto.split()).upper().strip('.')
        
        else:
            print(f"⚠️ Alerta: Fallo de parsing en: {linea}. Conductor/Territorio no encontrado.")
            lugar_bruto, conductor, terr_final = cuerpo, "ERROR", "T.XX"
        
        # Lógica de Hora mejorada
        hora_final = data['Hora']
        periodo = data.get('Periodo')

        if periodo:
            hora_final += periodo
        else:
            if data['Dia'].lower() in ['lunes', 'martes', 'miércoles', 'jueves', 'viernes'] and '7:00' in hora_final:
                hora_final += 'pm'
            elif data['Dia'].lower() not in ['sábado', 'domingo']:
                hora_final += 'am'

        datos_limpios.append({
            'DÍA': data['Dia'].capitalize(),
            'HORA': hora_final,
            'CONDUCTOR': conductor,
            'LUGAR_BRUTO': lugar_bruto, 
            'TERRITORIO': terr_final
        })

        print(f"✅ Procesado: {datos_limpios[-1]}")

df = pd.DataFrame(datos_limpios)
df['LUGAR DE PREDICACION'] = df['LUGAR_BRUTO']
df['MÉTODO DE PREDICACIÓN'] = 'PREDICACION Y NO EN CASA P.'
df['ID_ASIGNACION'] = range(1, len(df) + 1) 

columnas_finales = ['ID_ASIGNACION', 'DÍA', 'HORA', 'CONDUCTOR', 'LUGAR DE PREDICACION', 'MÉTODO DE PREDICACIÓN', 'TERRITORIO']
df_final = df[columnas_finales]

# 3. EXPORTACIÓN A LA PLANTILLA EXISTENTE.

PLANTILLA_FILE = "./template/template.xlsx" 

# El nombre será por semanas. Es decir, me pedirá fecha de inicio de semana y se hará un nombre como: ROL-(27/10-02/11)-2025.xlsx
fecha_inicio_str = input("\nIngresa la fecha de inicio de la semana (DD/MM/AAAA): ")
try:
    fecha_inicio = datetime.strptime(fecha_inicio_str, "%d/%m/%Y")
    fecha_fin = fecha_inicio + pd.Timedelta(days=6)
except ValueError:
    print("❌ ERROR: Formato de fecha inválido. Usa DD/MM/AAAA.")
    sys.exit(1)
HOJA_DATOS = 'Datos_Crudos'

# Configuración de Idioma para Fechas
try:
    locale.setlocale(locale.LC_TIME, 'es-ES')
except locale.Error:
    print("⚠️ Alerta: No se pudo configurar el idioma a español. Las fechas podrían aparecer en inglés.")

try:
    print("\nIniciando exportación a Excel...")
    wb = load_workbook(PLANTILLA_FILE)

    if HOJA_DATOS in wb.sheetnames:
        del wb[HOJA_DATOS]
    
    ws_datos = wb.create_sheet(HOJA_DATOS)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_final, header=True, index=False)):
        ws_datos.append(row)

    if fecha_inicio.strftime('%B') == fecha_fin.strftime('%B'):
        # Mismo mes: DEL 27 al 2 de Noviembre de 2025
        rango_fechas_texto = f"Del {fecha_inicio.day} al {fecha_fin.day} de {fecha_inicio.strftime('%B')}"
    else:
        # Meses diferentes: DEL 27 de Octubre al 2 de Noviembre de 2025
        rango_fechas_texto = f"Del {fecha_inicio.day} de {fecha_inicio.strftime('%B')} al {fecha_fin.day} de {fecha_fin.strftime('%B')}"

    ws_datos['A' + str(len(df_final) + 3)] = rango_fechas_texto
    OUTPUT_FILE = f"./output/ROL-({rango_fechas_texto} {fecha_inicio.year}).xlsx"
        
    wb.save(OUTPUT_FILE)
    
    print(f"\n✅ Proceso completado exitosamente.")
    print(f"Archivo generado: {OUTPUT_FILE}")

except FileNotFoundError:
    print(f"\n❌ ERROR: No se encontró el archivo de plantilla: {PLANTILLA_FILE}")
except Exception as e:
    print(f"\n❌ ERROR INESPERADO: {e}")